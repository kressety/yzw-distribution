import numpy as np
from openpyxl.reader.excel import load_workbook
from collections.abc import Iterable


def yzw_distribution(
    total: int, 
    zero_months: list[int], 
    higher_months: list[int], 
    performance: float, 
    noise: float = 0.05, 
    purging: bool = False
) -> list[int]:
    """志伟分布（套皮多项分布）：研究如何将总体值按指定规则分配到12个月份中。

    :param total: 总体值（总客流量）。
    :param zero_months: 客流量为0的月份。
    :param higher_months: 客流量较高的月份。
    :param performance: 业绩参数，用于控制客流量较高的月份高于其他月份的程度。如果该值小于1，则会起到相反的效果。
    :param noise: 扰动（噪声），使得数据分布显得更加不规律。当设置了较高的performance时，应将此值调低。
    :param purging: 如果将此项置为True，则客流量为0的月份不受扰动，保持为0。
    :return: 客流量实际分布。
    """
    normal_months_number = 12 - (len(zero_months) + len(higher_months))
    weight = normal_months_number + performance * len(higher_months)
    probs_per_unit = 1.0 / weight

    probs_list = []
    for this_month in range(12):
        if this_month in zero_months:
            probs_list.append(0)
        elif this_month in higher_months:
            probs_list.append(probs_per_unit * performance)
        else:
            probs_list.append(probs_per_unit * 1.0)
    probs_list = np.array(probs_list)

    noise_list = np.random.uniform(-noise, noise, size=len(probs_list))
    if purging:
        noise_list[probs_list == 0] = 0
    probs_list_with_noise = np.array(probs_list) + noise_list # 施加扰动
    probs_list_with_noise = np.clip(probs_list_with_noise, 0, 1) # 裁剪，确保每个概率都在[0,1)区间内
    probs_list_with_noise = probs_list_with_noise / probs_list_with_noise.sum() # 归一化

    return np.random.multinomial(total, probs_list_with_noise)


def write_excel(
    obj: list, 
    file_path: str, 
    sheet_name: str, 
    cells: list
):
    assert isinstance(obj, Iterable), 'Data should be iterable, such as list[] or tuple(). '
    assert isinstance(cells, Iterable), 'Cells should be iterable, such as list[] or tuple(). '
    assert len(obj) == len(cells), 'Number of data should equal to number of cells. '

    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]
    for i, cell in enumerate(cells):
        col = ord(cell[0].upper()) - 64
        row = int(cell[1:])
        worksheet.cell(row, col).value = obj[i]
    workbook.save(file_path)


if __name__ == '__main__':
    # 你的代码
    pass
